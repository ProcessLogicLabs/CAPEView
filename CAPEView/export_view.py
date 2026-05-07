"""Pure-Python writers for "export current view" — used by the toolbar button
on each table tab plus the Ctrl+E shortcut.

The table-walking glue (reading rows + tint colors from the QTableWidget) lives
in ``views.table_view.SQLTableView.export_current_view``; this module just
takes plain-Python inputs (list of header strings, list of row-of-strings,
list of optional row-tint hex colors like ``"#F5D2D7"``) so it's testable
without a QApplication.
"""

from __future__ import annotations

import csv
from pathlib import Path


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    """Write headers + rows to a UTF-8 CSV (BOM-prefixed for Excel
    compatibility). Currency / flag / date columns use the same display
    strings the user sees on screen."""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def write_xlsx(
    path: Path,
    sheet_title: str,
    headers: list[str],
    rows: list[list[str]],
    row_colors: list[str | None] | None = None,
) -> None:
    """Write headers + rows to an xlsx with cyan header styling, frozen header
    row, auto-sized columns, and per-row tint fills (matching the on-screen
    urgency colors). ``row_colors`` is a list aligned with ``rows``; each
    element is a ``#RRGGBB`` hex string or ``None`` for no tint."""
    # openpyxl is already a project dependency for workbook_export
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    # Excel sheet names are 31 chars max and must avoid : \ / ? * [ ]
    safe_title = "".join(c for c in (sheet_title or "Sheet1") if c not in r":\/?*[]")
    ws.title = (safe_title[:31] or "Sheet1")

    header_fill = PatternFill(start_color="FF4E8C9B", end_color="FF4E8C9B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    row_colors = row_colors or [None] * len(rows)
    for r_idx, (row, hex_color) in enumerate(zip(rows, row_colors), start=2):
        fill = None
        if hex_color:
            argb = "FF" + hex_color.lstrip("#").upper()
            fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill is not None:
                cell.fill = fill

    # Auto-fit columns based on visible content
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            n = len(str(v))
            if n > max_len:
                max_len = n
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    ws.freeze_panes = "A2"
    wb.save(str(path))
