"""Regenerate the legacy CAPE ESTIMATE workbook from cape.db.

Generates a workbook with the same core tabs downstream consumers expect:
  - Entry Count
  - Main Report
  - Claim details
  - Parameters

Pivot tabs are not regenerated as live PivotTables; instead, the underlying
ranges are written so an Excel user can refresh / rebuild pivots from the
generated sheets.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from CAPEView import cape_database as db
from CAPEView.version import get_version

logger = logging.getLogger(__name__)


ENTRY_COUNT_HEADERS = [
    "Entry Summary Number", "DIV", "CAPE Phase 1 Eligible", "SELF FILER",
    "ACE ACCOUNT", "ACH DETAILS IN ACE", "4811 CLIENT", "PSC FOR 4811",
    "CAPE CLAIM NUMBER", "FILING STATUS", "CAPE ERROR DETAIL",
    "Entry Type Code", "Importer Number", "Importer Name", "Port of Entry Code",
    "Entry Date", "Entry Summary Date", "Initial Entry Summary Create Date",
    "Reconciliation Indicator", "Control Status", "Post Summary Correction Indicator",
    "Liquidation Date", "Liquidation Status", "Final Liquidation Date (LIQ +180)",
    "Finally Liquidated", "CAPE LIQ Deadline (LIQ + 80)", "Protest Number",
    "Protest Status", "Review Team Number", "Country of Origin Code",
    "Country of Export Code", "Total Liquidated Duty Amount",
]

MAIN_REPORT_HEADERS = [
    "Entry Summary Number", "CAPE PHASE 1 ELIGIBLE", "CLAIM NUMBER", "CLAIM ERROR",
    "Entry Type Code", "Importer Number", "Importer Name", "Port of Entry Code",
    "Entry Date", "Entry Summary Date", "Initial Entry Summary Create Date",
    "Cape Phase 1 Eligible", "Reconciliation Indicator", "Control Status",
    "Post Summary Correction Indicator", "Liquidation Date", "Liquidation Status",
    "Final Liquidation", "Protest Number", "Protest Status",
    "Entry Summary Line Number", "Review Team Number", "Country of Origin Code",
    "Country of Export Code", "Manufacturer ID", "Foreign Exporter ID",
    "Line SPI Code", "Tariff Ordinal Number", "HTS Number - Full",
    "Line Tariff Goods Value Amount", "Line Tariff Duty Amount",
]

CLAIM_HEADERS = ["ENTRY_NUMBER", "CLAIM_NUMBER", "STATUS", "ERROR_DESCRIPTION"]


HEADER_FILL = PatternFill(start_color="4E8C9B", end_color="4E8C9B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)


def _style_header_row(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _write_entry_count(ws, conn):
    ws.append(ENTRY_COUNT_HEADERS)
    _style_header_row(ws, len(ENTRY_COUNT_HEADERS))

    rows = conn.execute(
        """SELECT e.entry_summary_number, e.div, e.cape_phase1_eligible,
                  i.self_filer, i.ace_account, i.ach_details_in_ace,
                  i.is_4811_client, i.psc_for_4811,
                  c.claim_number, c.status, c.error_description,
                  e.entry_type_code, e.importer_number, e.importer_name,
                  e.port_of_entry_code, e.entry_date, e.entry_summary_date,
                  e.initial_es_create_date, e.reconciliation_indicator,
                  e.control_status, e.psc_indicator, e.liquidation_date,
                  e.liquidation_status, e.final_liquidation_date,
                  CASE WHEN e.final_liquidation_date IS NOT NULL
                       AND date(e.final_liquidation_date) <= date('now')
                       THEN 'Y' ELSE 'N' END,
                  e.cape_liq_deadline, e.protest_number, e.protest_status,
                  e.review_team_number, e.country_of_origin_code,
                  e.country_of_export_code, e.total_liquidated_duty
           FROM entries e
           LEFT JOIN importer_status i ON i.importer_number = e.importer_number
           LEFT JOIN claims c ON c.entry_summary_number = e.entry_summary_number
           ORDER BY e.cape_liq_deadline IS NULL, e.cape_liq_deadline ASC""",
    )
    count = 0
    for r in rows:
        ws.append(list(r))
        count += 1
    return count


def _write_main_report(ws, conn):
    ws.append(MAIN_REPORT_HEADERS)
    _style_header_row(ws, len(MAIN_REPORT_HEADERS))

    rows = conn.execute(
        """SELECT e.entry_summary_number, e.cape_phase1_eligible,
                  c.claim_number, c.error_description,
                  e.entry_type_code, e.importer_number, e.importer_name,
                  e.port_of_entry_code, e.entry_date, e.entry_summary_date,
                  e.initial_es_create_date, e.cape_phase1_eligible,
                  e.reconciliation_indicator, e.control_status, e.psc_indicator,
                  e.liquidation_date, e.liquidation_status, e.final_liquidation_date,
                  e.protest_number, e.protest_status,
                  l.line_number, e.review_team_number,
                  COALESCE(l.country_of_origin_code, e.country_of_origin_code),
                  COALESCE(l.country_of_export_code, e.country_of_export_code),
                  l.manufacturer_id, l.foreign_exporter_id, l.line_spi_code,
                  l.tariff_ordinal, l.hts_number,
                  l.line_tariff_goods_value, l.line_tariff_duty
           FROM entries e
           LEFT JOIN entry_lines l ON l.entry_summary_number = e.entry_summary_number
           LEFT JOIN claims c ON c.entry_summary_number = e.entry_summary_number
           ORDER BY e.entry_summary_number, l.line_number, l.tariff_ordinal""",
    )
    count = 0
    for r in rows:
        ws.append(list(r))
        count += 1
    return count


def _write_claim_details(ws, conn):
    ws.append(CLAIM_HEADERS)
    _style_header_row(ws, len(CLAIM_HEADERS))
    rows = conn.execute(
        "SELECT entry_summary_number, claim_number, status, error_description "
        "FROM claims ORDER BY last_seen DESC"
    )
    count = 0
    for r in rows:
        ws.append(list(r))
        count += 1
    return count


def _write_parameters(ws, conn):
    ws["B1"] = "IEEPA DUTY PAID with LIQUIDATION DATE"
    ws["D1"] = "FOR OFFICIAL USE ONLY"
    ws["A4"] = "Report Parameters:"
    ws["A5"] = "Generated by CAPEView " + get_version()
    ws["A6"] = ("Generated at: "
                + datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds")
                + " UTC")
    ws["A8"] = f"Database: {db.resolve_db_path()}"


def export_cape_estimate(target: Path) -> dict:
    """Generate the workbook at ``target``. Returns row-count summary."""
    target = Path(target)
    target.parent.mkdir(parents=True, exist_ok=True)

    conn = db.connect()
    db.init_db(conn)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_entries = wb.create_sheet("Entry Count")
    ws_main = wb.create_sheet("Main Report")
    ws_claims = wb.create_sheet("Claim details")
    ws_params = wb.create_sheet("Parameters")

    entries_n = _write_entry_count(ws_entries, conn)
    main_n = _write_main_report(ws_main, conn)
    claims_n = _write_claim_details(ws_claims, conn)
    _write_parameters(ws_params, conn)

    wb.save(target)
    conn.close()

    return {"entries": entries_n, "entry_lines": main_n, "claims": claims_n, "path": str(target)}


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, required=True, help="Output xlsx path")
    args = parser.parse_args()
    summary = export_cape_estimate(args.out)
    print(summary)
