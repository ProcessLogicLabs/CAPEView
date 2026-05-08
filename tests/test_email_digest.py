"""Tests for the Compliance digest email module.

Outlook COM is mocked throughout — these tests run on any platform with
no pywin32 install required.
"""

from __future__ import annotations

from datetime import date, timedelta
from unittest.mock import MagicMock, patch

import pytest

from CAPEView import cape_database as db
from CAPEView import email_digest


# ---------------------------------------------------------------------------
# Fixtures

@pytest.fixture()
def seeded_db(tmp_path, monkeypatch):
    """Synthetic DB with two failed claims and one updated claim. The two
    failed claims have known cape_liq_deadline + duty values so summary
    aggregation is testable."""
    path = tmp_path / "cape.db"
    monkeypatch.setenv("CAPEVIEW_DB_PATH", str(path))
    conn = db.connect(path)
    db.init_db(conn)

    base = date.today()
    entries = [
        # esn,    importer_no,  importer,        div,         cape, liq_status,    liq, cape_liq, final_liq, total_duty
        ("E001", "11-1111111", "ACME INC",       "HOUSTON",   "Y", "Liquidated",
         base.isoformat(), (base + timedelta(days=10)).isoformat(),
         (base + timedelta(days=110)).isoformat(),  500.00),
        ("E002", "22-2222222", "WIDGETCO",       "LOS ANGELES","Y", "Liquidated",
         base.isoformat(), (base + timedelta(days=45)).isoformat(),
         (base + timedelta(days=145)).isoformat(),  300.00),
        ("E003", "33-3333333", "FRANKLIN INC",   "HOUSTON",   "Y", "Liquidated",
         base.isoformat(), (base + timedelta(days=80)).isoformat(),
         (base + timedelta(days=180)).isoformat(),  120.00),
    ]
    for esn, imp_no, imp_name, div_, cape, liq_status, liq, cape_liq, final_liq, duty in entries:
        conn.execute(
            "INSERT INTO entries (entry_summary_number, importer_number, importer_name, div, "
            " cape_phase1_eligible, liquidation_status, liquidation_date, cape_liq_deadline, "
            " final_liquidation_date, total_liquidated_duty, last_imported_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))",
            (esn, imp_no, imp_name, div_, cape, liq_status, liq, cape_liq, final_liq, duty),
        )

    # Two failed (E001, E003), one updated (E002)
    db.upsert_claims(conn, [
        {"entry_summary_number": "E001", "claim_number": "C001",
         "status": "Failed", "error_description": "UNABLE TO CALCULATE DUTY"},
        {"entry_summary_number": "E002", "claim_number": "C002",
         "status": "Entry Summary Updated", "error_description": ""},
        {"entry_summary_number": "E003", "claim_number": "C003",
         "status": "Failed", "error_description": "PROTEST ON ENTRY"},
    ])
    yield conn
    conn.close()


# ---------------------------------------------------------------------------
# build_digest

def test_build_digest_returns_only_currently_failed(seeded_db):
    digest = email_digest.build_digest(seeded_db)
    esns = [r[0] for r in digest["rows"]]
    assert set(esns) == {"E001", "E003"}
    assert "E002" not in esns


def test_build_digest_summary_stats(seeded_db):
    digest = email_digest.build_digest(seeded_db)
    s = digest["summary"]
    assert s["total_failed"] == 2
    # E001 = 500.00 + E003 = 120.00
    assert s["total_duty"] == pytest.approx(620.00)
    # E001 -> HOUSTON; E003 -> HOUSTON; both rows from HOUSTON
    assert s["by_div"] == {"HOUSTON": 2}


def test_build_digest_sorted_by_deadline_asc(seeded_db):
    digest = email_digest.build_digest(seeded_db)
    deadlines = [r[5] for r in digest["rows"]]
    # E001 deadline (today+10) is sooner than E003 (today+80)
    assert deadlines[0] < deadlines[1]


# ---------------------------------------------------------------------------
# render_html

def test_render_html_contains_summary_and_top10(seeded_db):
    digest = email_digest.build_digest(seeded_db)
    body = email_digest.render_html(digest, "v1.2.3")
    assert "Compliance digest" in body
    assert "Total failed claims" in body
    assert "v1.2.3" in body
    assert "Top 10" in body or "Top 10 most-urgent rejects" in body
    # Currency formatting in body
    assert "$620.00" in body
    # E001's importer should be HTML-escaped and present
    assert "ACME INC" in body


def test_render_html_escapes_user_strings():
    """Importer name / error description with HTML-special chars must be escaped."""
    digest = {
        "summary": {"total_failed": 1, "total_duty": 100.0,
                    "by_div": {"HOUSTON": 1}},
        "rows": [
            ("E001", "C001", "Failed",
             "<script>alert('xss')</script>",
             "HOUSTON", "2026-12-01", 100.0,
             "<b>Bobby Tables & Co</b>"),
        ],
    }
    body = email_digest.render_html(digest, "v0.0.1")
    # The raw <script> must NOT appear; escaped form must
    assert "<script>alert" not in body
    assert "&lt;script&gt;" in body
    # Importer name escaped
    assert "&lt;b&gt;Bobby Tables &amp; Co" in body


def test_render_html_handles_zero_failed():
    digest = {"summary": {"total_failed": 0, "total_duty": 0,
                          "by_div": {}}, "rows": []}
    body = email_digest.render_html(digest, "v0.0.1")
    assert "All clear" in body
    assert "Top 10" not in body


# ---------------------------------------------------------------------------
# write_attachment

def test_write_attachment_produces_readable_xlsx(seeded_db, tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    digest = email_digest.build_digest(seeded_db)
    target = tmp_path / "digest.xlsx"
    email_digest.write_attachment(digest, target)

    wb = load_workbook(target)
    ws = wb.active
    assert ws.title == "Compliance"
    # Header row matches HEADERS
    assert [c.value for c in ws[1]] == email_digest.HEADERS
    # Data rows = 2 failed claims
    assert ws.max_row == 1 + 2


# ---------------------------------------------------------------------------
# Outlook COM transport (mocked)

def _make_disabled_settings():
    s = MagicMock()
    # Any key returns the second arg (the default) — disabled, empty list
    s.get.side_effect = lambda key, default=None: default
    s.get.return_value = False
    return s


def _make_settings(*, enabled=True, recipients=None):
    """Build a mock SettingsManager that returns ``enabled`` for
    ``email.enabled`` and ``recipients`` for ``email.recipients``."""
    s = MagicMock()

    def fake_get(key, default=None):
        if key == "email.enabled":
            return enabled
        if key == "email.recipients":
            return recipients if recipients is not None else []
        return default

    s.get.side_effect = fake_get
    return s


def test_send_is_noop_when_disabled(seeded_db):
    result = email_digest.send_compliance_digest(
        seeded_db, "v0.0.1", settings=_make_settings(enabled=False),
    )
    assert result == {"sent": False, "recipients": [], "rows": 0, "error": None}


def test_send_returns_error_when_no_recipients(seeded_db):
    """email.enabled=True, recipients empty, Outlook unreachable → graceful fail."""
    with patch.object(email_digest, "_get_current_user_email", return_value=""):
        result = email_digest.send_compliance_digest(
            seeded_db, "v0.0.1", settings=_make_settings(recipients=[]),
        )
    assert result["sent"] is False
    assert result["recipients"] == []
    assert "no recipients" in (result["error"] or "")


def test_resolve_recipients_uses_configured_list():
    """A non-empty recipients list wins over the self-fallback."""
    settings = _make_settings(recipients=["a@x.com", "b@x.com"])
    with patch.object(email_digest, "_get_current_user_email",
                      return_value="should-not-be-called@example.com"):
        addrs = email_digest._resolve_recipients(settings)
    assert addrs == ["a@x.com", "b@x.com"]


def test_resolve_recipients_falls_back_to_self_when_empty():
    settings = _make_settings(recipients=[])
    with patch.object(email_digest, "_get_current_user_email",
                      return_value="me@example.com"):
        addrs = email_digest._resolve_recipients(settings)
    assert addrs == ["me@example.com"]


def test_resolve_recipients_strips_blank_entries():
    settings = _make_settings(recipients=["a@x.com", "", "  ", "b@x.com"])
    with patch.object(email_digest, "_get_current_user_email",
                      return_value="ignored@example.com"):
        addrs = email_digest._resolve_recipients(settings)
    assert addrs == ["a@x.com", "b@x.com"]


def test_send_via_outlook_invokes_correct_calls(seeded_db, tmp_path):
    """Mock the module-level win32com reference and verify the call sequence."""
    fake_dispatch = MagicMock()
    fake_outlook = MagicMock()
    fake_mail = MagicMock()
    fake_dispatch.return_value = fake_outlook
    fake_outlook.CreateItem.return_value = fake_mail

    fake_module = MagicMock()
    fake_module.Dispatch = fake_dispatch

    # Pre-create attachment so the path exists when Outlook would attach it
    attachment = tmp_path / "att.xlsx"
    attachment.write_bytes(b"fake xlsx")

    # Patch the module-level reference directly — no sys.modules munging.
    with patch.object(email_digest, "_win32com_client", fake_module):
        email_digest._send_via_outlook(
            "Test Subject", "<p>body</p>", attachment, "user@example.com",
        )

    fake_outlook.CreateItem.assert_called_once_with(0)
    assert fake_mail.Subject == "Test Subject"
    assert fake_mail.HTMLBody == "<p>body</p>"
    assert fake_mail.To == "user@example.com"
    fake_mail.Attachments.Add.assert_called_once_with(str(attachment))
    fake_mail.Send.assert_called_once()


def test_send_via_outlook_raises_when_pywin32_missing(tmp_path):
    """If pywin32 isn't available the function raises rather than silently
    succeeding (the caller in send_compliance_digest catches and reports)."""
    attachment = tmp_path / "att.xlsx"
    attachment.write_bytes(b"x")
    with patch.object(email_digest, "_win32com_client", None):
        with pytest.raises(RuntimeError, match="pywin32 not installed"):
            email_digest._send_via_outlook(
                "S", "<p></p>", attachment, "u@example.com",
            )


def test_send_compliance_digest_end_to_end_self(seeded_db, tmp_path):
    """Empty recipients list + valid Outlook self → digest sends to self."""
    with patch.object(email_digest, "_get_current_user_email",
                      return_value="heath@example.com"), \
         patch.object(email_digest, "_send_via_outlook") as mock_send:
        result = email_digest.send_compliance_digest(
            seeded_db, "v0.0.1", settings=_make_settings(recipients=[]),
        )

    assert result["sent"] is True
    assert result["recipients"] == ["heath@example.com"]
    assert result["rows"] == 2
    assert result["error"] is None
    mock_send.assert_called_once()
    args, _ = mock_send.call_args
    subject, _html_body, attachment_path, to_address = args
    assert "Compliance digest" in subject
    assert "2 failed claims" in subject
    assert to_address == "heath@example.com"
    assert attachment_path.exists()


def test_send_compliance_digest_end_to_end_multi_recipient(seeded_db, tmp_path):
    """Configured recipients are joined into a semicolon-delimited Outlook To."""
    with patch.object(email_digest, "_get_current_user_email",
                      return_value="should-not-be-used@example.com"), \
         patch.object(email_digest, "_send_via_outlook") as mock_send:
        result = email_digest.send_compliance_digest(
            seeded_db, "v0.0.1",
            settings=_make_settings(recipients=["eric@example.com", "team@example.com"]),
        )

    assert result["sent"] is True
    assert result["recipients"] == ["eric@example.com", "team@example.com"]
    args, _ = mock_send.call_args
    _subject, _html_body, _attachment, to_address = args
    assert to_address == "eric@example.com; team@example.com"


# ---------------------------------------------------------------------------
# Ingest hook

def test_ingest_calls_digest_when_rows_changed(tmp_path, monkeypatch):
    """process_inbox should call _maybe_send_digest after a successful ingest."""
    from CAPEView import claims_csv_ingest

    # Isolate the DB
    db_path = tmp_path / "cape.db"
    monkeypatch.setenv("CAPEVIEW_DB_PATH", str(db_path))

    # Spy on the digest-send function
    captured = {"called": False, "rows_at_call": None}

    def fake_send(conn, version, settings=None):
        captured["called"] = True
        # Connection should be live; rows should reflect the just-ingested state
        captured["rows_at_call"] = conn.execute(
            "SELECT COUNT(*) FROM claims"
        ).fetchone()[0]
        return {"sent": False, "recipient": None, "rows": 0, "error": None}

    monkeypatch.setattr(email_digest, "send_compliance_digest", fake_send)

    # Drop a CSV in
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    csv_file = inbox / "claims.csv"
    csv_file.write_text(
        "ENTRY_NUMBER,CLAIM_NUMBER,STATUS,ERROR_DESCRIPTION\n"
        "60500009000,X1,Failed,UNABLE TO CALCULATE DUTY\n",
        encoding="utf-8",
    )
    summary = claims_csv_ingest.process_inbox(inbox)

    assert summary["inserted"] == 1
    assert captured["called"] is True
    assert captured["rows_at_call"] == 1


def test_ingest_skips_digest_when_no_compliance_change(tmp_path, monkeypatch):
    """Re-ingesting a CSV with the same Failed-status row → no digest.

    This is the spam-prevention guard: upsert_claims counts every existing
    row as 'updated' even when only last_seen got bumped, so the previous
    rule of 'fire when inserted+updated > 0' caused 12 emails for 12
    drops of the same file. The new rule fires only when audit_log shows
    a status transition or a new Failed claim arrived this cycle.
    """
    from CAPEView import claims_csv_ingest

    db_path = tmp_path / "cape.db"
    monkeypatch.setenv("CAPEVIEW_DB_PATH", str(db_path))

    captured = {"calls": 0}

    def fake_send(*args, **kwargs):
        captured["calls"] += 1
        return {"sent": False, "recipients": [], "rows": 0, "error": None}

    monkeypatch.setattr(email_digest, "send_compliance_digest", fake_send)

    inbox = tmp_path / "inbox"
    inbox.mkdir()
    csv_content = (
        "ENTRY_NUMBER,CLAIM_NUMBER,STATUS,ERROR_DESCRIPTION\n"
        "60500009000,X1,Failed,UNABLE TO CALCULATE DUTY\n"
    )

    # First ingest: brand-new Failed claim → digest fires
    (inbox / "first.csv").write_text(csv_content, encoding="utf-8")
    claims_csv_ingest.process_inbox(inbox)
    assert captured["calls"] == 1

    # Second ingest: same content, no real changes → digest does NOT fire
    (inbox / "second.csv").write_text(csv_content, encoding="utf-8")
    claims_csv_ingest.process_inbox(inbox)
    assert captured["calls"] == 1, "digest should not fire when nothing changed"


def test_ingest_calls_digest_on_status_transition(tmp_path, monkeypatch):
    """Status flip Failed → Updated produces an audit_log entry → digest fires."""
    from CAPEView import claims_csv_ingest

    db_path = tmp_path / "cape.db"
    monkeypatch.setenv("CAPEVIEW_DB_PATH", str(db_path))

    captured = {"calls": 0}

    def fake_send(*args, **kwargs):
        captured["calls"] += 1
        return {"sent": False, "recipients": [], "rows": 0, "error": None}

    monkeypatch.setattr(email_digest, "send_compliance_digest", fake_send)

    inbox = tmp_path / "inbox"
    inbox.mkdir()

    # Day 1: claim is Failed
    (inbox / "day1.csv").write_text(
        "ENTRY_NUMBER,CLAIM_NUMBER,STATUS,ERROR_DESCRIPTION\n"
        "60500009000,X1,Failed,UNABLE TO CALCULATE DUTY\n",
        encoding="utf-8",
    )
    claims_csv_ingest.process_inbox(inbox)
    assert captured["calls"] == 1

    # Day 2: claim flips to Entry Summary Updated → digest fires (transition out of Failed)
    (inbox / "day2.csv").write_text(
        "ENTRY_NUMBER,CLAIM_NUMBER,STATUS,ERROR_DESCRIPTION\n"
        "60500009000,X1,Entry Summary Updated,\n",
        encoding="utf-8",
    )
    claims_csv_ingest.process_inbox(inbox)
    assert captured["calls"] == 2


def test_ingest_skips_digest_when_no_rows(tmp_path, monkeypatch):
    """An empty inbox produces a summary with 0 rows; digest should NOT fire."""
    from CAPEView import claims_csv_ingest

    db_path = tmp_path / "cape.db"
    monkeypatch.setenv("CAPEVIEW_DB_PATH", str(db_path))

    captured = {"called": False}

    def fake_send(*args, **kwargs):
        captured["called"] = True
        return {"sent": False, "recipient": None, "rows": 0, "error": None}

    monkeypatch.setattr(email_digest, "send_compliance_digest", fake_send)

    inbox = tmp_path / "inbox"
    inbox.mkdir()
    summary = claims_csv_ingest.process_inbox(inbox)

    assert summary["inserted"] == 0
    assert captured["called"] is False
