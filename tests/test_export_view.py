"""Tests for export_view writers — pure Python, no Qt required."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from CAPEView import export_view


HEADERS = ["Entry Summary #", "Status", "Total Liq Duty", "ACE"]
ROWS = [
    ["60576072486", "Failed", "$26,352,732.57", "Y"],
    ["60576072809", "Entry Summary Updated", "", "N"],
    ["60525904516", "Failed", "$1,234.50", ""],
]
ROW_COLORS = ["#F5D2D7", None, "#F5D2D7"]


def test_write_csv_roundtrip(tmp_path):
    path = tmp_path / "out.csv"
    export_view.write_csv(path, HEADERS, ROWS)
    with open(path, encoding="utf-8-sig", newline="") as f:
        read = list(csv.reader(f))
    assert read[0] == HEADERS
    assert read[1] == ROWS[0]
    assert len(read) == 1 + len(ROWS)


def test_write_xlsx_basic(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    path = tmp_path / "out.xlsx"
    export_view.write_xlsx(path, "Entries", HEADERS, ROWS)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "Entries"
    assert ws.freeze_panes == "A2"
    assert [c.value for c in ws[1]] == HEADERS
    assert [c.value for c in ws[2]] == ROWS[0]
    assert ws.cell(row=1, column=1).font.bold is True


def test_write_xlsx_applies_row_tint(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    path = tmp_path / "out.xlsx"
    export_view.write_xlsx(path, "Entries", HEADERS, ROWS, row_colors=ROW_COLORS)

    wb = load_workbook(path)
    ws = wb.active
    # Row 2 (first data row) should have the tint; row 3 should not
    tinted = ws.cell(row=2, column=1).fill
    untinted = ws.cell(row=3, column=1).fill
    assert tinted.fgColor.rgb == "FFF5D2D7"
    # An unfilled cell has fill_type None or 'none'
    assert untinted.fill_type in (None, "none")


def test_write_xlsx_sanitizes_sheet_title(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    path = tmp_path / "out.xlsx"
    # Excel rejects : \ / ? * [ ] in sheet names; writer must strip them
    export_view.write_xlsx(path, "Tab/with*bad:chars", HEADERS, ROWS)
    wb = load_workbook(path)
    assert wb.active.title == "Tabwithbadchars"


def test_write_xlsx_truncates_long_sheet_title(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    path = tmp_path / "out.xlsx"
    long = "A" * 50
    export_view.write_xlsx(path, long, HEADERS, ROWS)
    wb = load_workbook(path)
    assert len(wb.active.title) == 31


def test_write_xlsx_handles_empty_rows(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    path = tmp_path / "out.xlsx"
    export_view.write_xlsx(path, "Empty", HEADERS, [])
    wb = load_workbook(path)
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    # Only the header row exists
    assert ws.max_row == 1
