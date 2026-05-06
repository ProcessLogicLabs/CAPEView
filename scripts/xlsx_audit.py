"""Forensic audit of the legacy CAPE ESTIMATE workbook.

Reads the xlsx package directly via zipfile + XML parsing for the workbook-level
metadata (sheets, names, pivots, queries, macros, charts, slicers), and uses
openpyxl in **read-only** mode for sampled formula scanning. This avoids the
non-read-only loader's full-cell-graph hydration, which OOMs on workbooks with
198K+ formula rows like Main Report.

Usage:
  python scripts/xlsx_audit.py [--xlsx PATH] [--max-formula-rows N]
"""

from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter
from pathlib import Path

import openpyxl

DEFAULT_XLSX = (
    Path(__file__).resolve().parents[2]
    / "CAPEApp"
    / "Resources"
    / "CAPE ESTIMATE with LIQUIDATION DATE 20260415.xlsx"
)

NS = {
    "x":  "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r":  "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rd": "http://schemas.openxmlformats.org/package/2006/relationships",
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
}


def section(title: str):
    print()
    print("=" * 78)
    print(title)
    print("=" * 78)


# ---------------------------------------------------------------------------
# Workbook-level metadata via direct XML

def load_xml(z: zipfile.ZipFile, name: str):
    with z.open(name) as f:
        return ET.fromstring(f.read())


def list_zip(xlsx_path: Path):
    section("PACKAGE CONTENTS (xlsx ZIP listing)")
    with zipfile.ZipFile(xlsx_path) as z:
        names = sorted(z.namelist())
    interesting = ["pivotTable", "pivotCache", "slicer", "/charts/", "queryTable",
                   "connection", "vbaProject", "customXml", "table"]
    for n in names:
        flag = next((f for f in interesting if f.lower() in n.lower()), "")
        marker = "  <-- " + flag if flag else ""
        if flag:
            print(f"  {n}{marker}")
    return names


def audit_sheets(xlsx_path: Path):
    section("SHEETS — names, visibility, sheetIds")
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = load_xml(z, "xl/workbook.xml")
        sheets = wb_xml.find("x:sheets", NS)
        for i, s in enumerate(sheets.findall("x:sheet", NS), 1):
            print(f"  [{i:2}]  name={s.get('name')!r:30}  "
                  f"sheetId={s.get('sheetId')}  state={s.get('state', 'visible')}  "
                  f"r:id={s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')}")


def audit_defined_names_xml(xlsx_path: Path):
    section("DEFINED NAMES (named ranges, print areas, table refs)")
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = load_xml(z, "xl/workbook.xml")
        names = wb_xml.find("x:definedNames", NS)
        if names is None:
            print("  (none)")
            return
        for n in names.findall("x:definedName", NS):
            nm = n.get("name")
            local = n.get("localSheetId")
            print(f"  {nm!r:35}  localSheetId={local!r:5}  ->  {(n.text or '').strip()[:140]}")


def audit_connections_queries(xlsx_path: Path):
    section("EXTERNAL CONNECTIONS / POWER QUERY")
    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()
        targets = [n for n in names if n.startswith("xl/connections")
                                    or n.startswith("xl/queryTables/")
                                    or n.startswith("xl/customXml/")]
        if not targets:
            print("  (none)")
            return
        for tgt in targets:
            print(f"\n  {tgt}")
            with z.open(tgt) as f:
                txt = f.read().decode(errors="replace")[:1000]
                print("  " + txt.replace("\n", "\n  "))


def audit_tables(xlsx_path: Path):
    section("STRUCTURED TABLES (xl/tables/*.xml)")
    with zipfile.ZipFile(xlsx_path) as z:
        names = [n for n in z.namelist() if n.startswith("xl/tables/")]
        if not names:
            print("  (none)")
            return
        for n in sorted(names):
            try:
                root = load_xml(z, n)
                print(f"  {n}  name={root.get('name')!r}  ref={root.get('ref')}  "
                      f"displayName={root.get('displayName')!r}")
            except Exception as e:
                print(f"  {n}: parse error {e}")


def audit_pivots(xlsx_path: Path):
    section("PIVOT TABLE DEFINITIONS")
    with zipfile.ZipFile(xlsx_path) as z:
        pv_files = [n for n in z.namelist() if "pivotTable" in n and n.endswith(".xml")]
        if not pv_files:
            print("  (none)")
            return
        for pv_file in sorted(pv_files):
            try:
                root = load_xml(z, pv_file)
            except Exception as e:
                print(f"\n  {pv_file}: parse error {e}")
                continue
            print(f"\n  {pv_file}")
            print(f"    name      = {root.get('name')}")
            print(f"    cacheId   = {root.get('cacheId')}")
            print(f"    rowGrandTotals = {root.get('rowGrandTotals')}, "
                  f"colGrandTotals = {root.get('colGrandTotals')}")

            df = root.find("x:dataFields", NS)
            if df is not None:
                print(f"    data fields ({len(df)}):")
                for d in df.findall("x:dataField", NS):
                    print(f"      name={d.get('name')!r:30}  "
                          f"fld={d.get('fld')}  subtotal={d.get('subtotal')}")
            for axis_label, axis_tag in (("row", "rowFields"), ("col", "colFields"),
                                         ("page (filter)", "pageFields")):
                el = root.find(f"x:{axis_tag}", NS)
                if el is None:
                    continue
                tag = "x:field" if axis_tag != "pageFields" else "x:pageField"
                kids = list(el.findall(tag, NS))
                attrs = [{"x": k.get("x"), "fld": k.get("fld"), "hier": k.get("hier")} for k in kids]
                print(f"    {axis_label} fields ({len(kids)}): {attrs}")


def audit_pivot_caches(xlsx_path: Path):
    section("PIVOT CACHE FIELDS (the pivot's source columns)")
    with zipfile.ZipFile(xlsx_path) as z:
        cf_files = [n for n in z.namelist() if "pivotCacheDefinition" in n]
        if not cf_files:
            print("  (none)")
            return
        for cf_file in sorted(cf_files):
            try:
                root = load_xml(z, cf_file)
            except Exception as e:
                print(f"\n  {cf_file}: parse error {e}")
                continue
            cs = root.find("x:cacheSource", NS)
            wsrc = cs.find("x:worksheetSource", NS) if cs is not None else None
            sheet = wsrc.get("sheet") if wsrc is not None else None
            ref = wsrc.get("ref") if wsrc is not None else None
            named = wsrc.get("name") if wsrc is not None else None
            print(f"\n  {cf_file}")
            print(f"    cacheSource type={cs.get('type') if cs is not None else None}  "
                  f"sheet={sheet!r}  ref={ref}  name={named!r}")
            cfields = root.find("x:cacheFields", NS)
            if cfields is not None:
                for i, cf in enumerate(cfields.findall("x:cacheField", NS)):
                    print(f"    cacheField[{i}]  name={cf.get('name')!r:36}  "
                          f"numFmt={cf.get('numFmtId')}")


def audit_slicers_charts_macros(xlsx_path: Path):
    section("SLICERS / CHARTS / MACROS")
    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()
        groups = {
            "Slicer caches":        [n for n in names if "slicerCache" in n.lower()],
            "Slicers":              [n for n in names if "slicers" in n.lower() and "cache" not in n.lower()],
            "Charts":               [n for n in names if "/charts/" in n.replace("\\", "/").lower()],
            "Macros (vbaProject)":  [n for n in names if "vbaProject" in n],
            "External links":       [n for n in names if "externalLinks" in n],
            "Drawings":             [n for n in names if "/drawings/" in n.replace("\\", "/")],
        }
        for label, items in groups.items():
            print(f"\n  {label}: {len(items)}")
            for n in sorted(items):
                print(f"    {n}")


def audit_conditional_formatting_xml(xlsx_path: Path):
    """Pull conditional format rules out of each sheet xml directly."""
    section("CONDITIONAL FORMATTING (per sheet)")
    with zipfile.ZipFile(xlsx_path) as z:
        sheet_files = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
        for sf in sorted(sheet_files):
            try:
                root = load_xml(z, sf)
            except Exception as e:
                print(f"  {sf}: parse error {e}")
                continue
            cfs = root.findall("x:conditionalFormatting", NS)
            if not cfs:
                continue
            print(f"\n  {sf}")
            for cf in cfs:
                sqref = cf.get("sqref")
                rules = cf.findall("x:cfRule", NS)
                for r in rules:
                    rtype = r.get("type")
                    op = r.get("operator")
                    pri = r.get("priority")
                    formula_el = r.find("x:formula", NS)
                    formula = formula_el.text if formula_el is not None else None
                    print(f"    sqref={sqref}  type={rtype}  op={op}  priority={pri}  formula={formula}")


def audit_data_validations_xml(xlsx_path: Path):
    section("DATA VALIDATIONS (per sheet)")
    with zipfile.ZipFile(xlsx_path) as z:
        sheet_files = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
        for sf in sorted(sheet_files):
            try:
                root = load_xml(z, sf)
            except Exception:
                continue
            dvs_el = root.find("x:dataValidations", NS)
            if dvs_el is None:
                continue
            for dv in dvs_el.findall("x:dataValidation", NS):
                print(f"  {sf}  ranges={dv.get('sqref')}  type={dv.get('type')}  "
                      f"op={dv.get('operator')}  f1={[f.text for f in dv.findall('x:formula1', NS)]}  "
                      f"f2={[f.text for f in dv.findall('x:formula2', NS)]}")


def audit_autofilters_xml(xlsx_path: Path):
    section("AUTOFILTER RANGES")
    with zipfile.ZipFile(xlsx_path) as z:
        sheet_files = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
        for sf in sorted(sheet_files):
            try:
                root = load_xml(z, sf)
            except Exception:
                continue
            af = root.find("x:autoFilter", NS)
            if af is None:
                continue
            print(f"  {sf}  range={af.get('ref')}")


# ---------------------------------------------------------------------------
# Sampled formula scan via openpyxl read_only

def audit_formulas_sampled(xlsx_path: Path, max_rows: int):
    section(f"FORMULAS — sampled from rows 1..{max_rows} (read-only loader)")
    wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    for s in wb_ro.sheetnames:
        ws = wb_ro[s]
        shapes: Counter = Counter()
        first_loc: dict[str, str] = {}
        cnt = 0
        for row in ws.iter_rows(min_row=1, max_row=max_rows):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cnt += 1
                    shape = re.sub(r"\$?\d+", "<R>", v)
                    shapes[shape] += 1
                    first_loc.setdefault(shape, cell.coordinate)
        print(f"\n  {s!r}  ({cnt} formula cells in first {max_rows} rows, "
              f"{len(shapes)} unique shapes)")
        for shape, count in shapes.most_common(20):
            print(f"    x{count:<5}  first@{first_loc[shape]:<6}  =>  {shape[:140]}")
    wb_ro.close()


# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--max-formula-rows", type=int, default=60)
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Workbook not found: {args.xlsx}")

    print(f"Auditing: {args.xlsx}")

    list_zip(args.xlsx)
    audit_sheets(args.xlsx)
    audit_defined_names_xml(args.xlsx)
    audit_tables(args.xlsx)
    audit_autofilters_xml(args.xlsx)
    audit_data_validations_xml(args.xlsx)
    audit_conditional_formatting_xml(args.xlsx)
    audit_pivots(args.xlsx)
    audit_pivot_caches(args.xlsx)
    audit_slicers_charts_macros(args.xlsx)
    audit_connections_queries(args.xlsx)
    audit_formulas_sampled(args.xlsx, args.max_formula_rows)

    print("\nAudit complete.")


if __name__ == "__main__":
    main()
