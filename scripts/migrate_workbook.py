"""One-time migration: load the legacy CAPE ESTIMATE workbook into cape.db.

Reads three sheets:
  - Entry Count   -> entries
  - Main Report   -> entry_lines
  - Claim details -> claims

Usage:
  python scripts/migrate_workbook.py [--xlsx PATH] [--db PATH]

Defaults:
  --xlsx  ../CAPEApp/Resources/CAPE ESTIMATE with LIQUIDATION DATE 20260415.xlsx
  --db    resolved by cape_database.resolve_db_path() (env CAPEVIEW_DB_PATH wins)
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

# Allow running as a script without installing the package.
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from CAPEView import cape_database as db  # noqa: E402

DEFAULT_XLSX = (
    Path(__file__).resolve().parents[2]
    / "CAPEApp"
    / "Resources"
    / "CAPE ESTIMATE with LIQUIDATION DATE 20260415.xlsx"
)


def _iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)


def _to_int_flag(v) -> int | None:
    if v is None or v == "":
        return None
    s = str(v).strip().upper()
    if s in ("Y", "YES", "TRUE", "1"):
        return 1
    if s in ("N", "NO", "FALSE", "0"):
        return 0
    return None


def _f(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_entry_count(ws) -> list[dict]:
    """Sheet columns A..AF (32 cols). Row 1 is the header."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append(
            {
                "entry_summary_number": str(row[0]).strip() if row[0] is not None else None,
                "div": row[1],
                "cape_phase1_eligible": row[2],
                "entry_type_code": row[11],
                "importer_number": row[12],
                "importer_name": row[13],
                "port_of_entry_code": row[14],
                "entry_date": _iso_date(row[15]),
                "entry_summary_date": _iso_date(row[16]),
                "initial_es_create_date": _iso_date(row[17]),
                "reconciliation_indicator": row[18],
                "control_status": row[19],
                "psc_indicator": row[20],
                "liquidation_date": _iso_date(row[21]),
                "liquidation_status": row[22],
                "final_liquidation_date": _iso_date(row[23]),
                # row[24] = "Finally Liquidated" (Y/N flag) — derived, not stored.
                "cape_liq_deadline": _iso_date(row[25]),
                "protest_number": row[26],
                "protest_status": row[27],
                "review_team_number": row[28],
                "country_of_origin_code": row[29],
                "country_of_export_code": row[30],
                "total_liquidated_duty": _f(row[31]),
            }
        )
    return rows


def load_main_report(ws) -> list[dict]:
    """Main Report has 41 columns; we map a subset into entry_lines."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            line_number = int(row[20]) if row[20] is not None else None
        except (TypeError, ValueError):
            line_number = None
        try:
            tariff_ord = int(row[30]) if row[30] is not None else None
        except (TypeError, ValueError):
            tariff_ord = None
        if line_number is None or tariff_ord is None:
            continue
        rows.append(
            {
                "entry_summary_number": str(row[0]).strip(),
                "line_number": line_number,
                "tariff_ordinal": tariff_ord,
                "hts_number": row[31],
                "line_tariff_goods_value": _f(row[38]),
                "line_tariff_duty": _f(row[39]),
                "manufacturer_id": row[24],
                "foreign_exporter_id": row[25],
                "line_spi_code": row[26],
                "country_of_origin_code": row[22],
                "country_of_export_code": row[23],
            }
        )
    return rows


def load_claim_details(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        rows.append(
            {
                "entry_summary_number": str(row[0]).strip(),
                "claim_number": str(row[1]).strip(),
                "status": row[2],
                "error_description": row[3],
            }
        )
    return rows


def insert_entry_lines(conn, rows: list[dict]):
    """Bulk insert entry_lines (no per-row upsert — this is a one-shot migration)."""
    cols = [
        "entry_summary_number", "line_number", "tariff_ordinal", "hts_number",
        "line_tariff_goods_value", "line_tariff_duty", "manufacturer_id",
        "foreign_exporter_id", "line_spi_code", "country_of_origin_code",
        "country_of_export_code",
    ]
    placeholders = ", ".join(["?"] * len(cols))
    sql = f"INSERT OR REPLACE INTO entry_lines ({', '.join(cols)}) VALUES ({placeholders})"
    with db.transaction(conn):
        conn.executemany(sql, [[r.get(c) for c in cols] for r in rows])


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Source workbook (default: {DEFAULT_XLSX})")
    parser.add_argument("--db", type=Path, default=None,
                        help="Target SQLite path (default: cape_database.resolve_db_path())")
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"ERROR: workbook not found: {args.xlsx}")

    print(f"Loading workbook: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    print("Initialising database schema...")
    conn = db.connect(args.db) if args.db else db.connect()
    db.init_db(conn)

    started = db.now_iso()

    if "Entry Count" in wb.sheetnames:
        print("Reading 'Entry Count'...")
        entries = load_entry_count(wb["Entry Count"])
        print(f"  {len(entries)} entry rows")
        ins, upd = db.upsert_entries(conn, entries)
        print(f"  inserted={ins} updated={upd}")
        db.record_import_run(conn, "migrate_workbook:entries",
                             str(args.xlsx), ins, upd, started)

    if "Main Report" in wb.sheetnames:
        print("Reading 'Main Report'...")
        lines = load_main_report(wb["Main Report"])
        print(f"  {len(lines)} entry-line rows")
        insert_entry_lines(conn, lines)
        db.record_import_run(conn, "migrate_workbook:entry_lines",
                             str(args.xlsx), len(lines), 0, started)

    if "Claim details" in wb.sheetnames:
        print("Reading 'Claim details'...")
        claims = load_claim_details(wb["Claim details"])
        print(f"  {len(claims)} claim rows")
        ins, upd = db.upsert_claims(conn, claims)
        print(f"  inserted={ins} updated={upd}")
        db.record_import_run(conn, "migrate_workbook:claims",
                             str(args.xlsx), ins, upd, started)

    print("Migration complete.")


if __name__ == "__main__":
    main()
